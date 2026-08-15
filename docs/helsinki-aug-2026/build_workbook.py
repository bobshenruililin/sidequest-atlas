#!/usr/bin/env python3
"""Helsinki 16–20 Aug 2026 workbook. Hours/prices: verify on the day. Do not invent."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "Helsinki-16-20-Aug-2026.xlsx"
ACCESSED = "2026-08-15"

NAVY, GOLD, CREAM, TEAL, CORAL = "1B3A4B", "C4A35A", "F7F1E3", "2A6F6F", "C45C26"
PALE, MUST, BOOK, WALK, SKIP = "E8EEF0", "F4D6C6", "D4E8D4", "FFF3C4", "E8D5D5"

header_fill = PatternFill("solid", fgColor=NAVY)
gold_fill = PatternFill("solid", fgColor=GOLD)
cream_fill = PatternFill("solid", fgColor=CREAM)
pale_fill = PatternFill("solid", fgColor=PALE)
must_fill = PatternFill("solid", fgColor=MUST)
book_fill = PatternFill("solid", fgColor=BOOK)
walk_fill = PatternFill("solid", fgColor=WALK)
skip_fill = PatternFill("solid", fgColor=SKIP)
open_fill = PatternFill("solid", fgColor="C8E6C9")
closed_fill = PatternFill("solid", fgColor="FFCDD2")
maybe_fill = PatternFill("solid", fgColor="FFE082")
white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_font = Font(name="Calibri", bold=True, color=NAVY, size=16)
section_font = Font(name="Calibri", bold=True, color=NAVY, size=13)
body = Font(name="Calibri", size=11)
link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def paint(ws, start=2):
    for row in ws.iter_rows(min_row=start, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            c.font = link_font if c.hyperlink else body
            c.alignment = wrap
            c.border = thin
        current = ws.row_dimensions[row[0].row].height
        ws.row_dimensions[row[0].row].height = max(current or 18, 42)


def href(ws, r, c, url, label=None):
    cell = ws.cell(r, c, label or url)
    if url and str(url).startswith("http"):
        cell.hyperlink = url
        cell.font = link_font
    return cell


def add_table(ws, headers, rows, widths, tab=None, row_height=48):
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    style_header(ws, len(headers))
    paint(ws)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = row_height
    autosize(ws, widths)
    if tab:
        ws.sheet_properties.tabColor = tab
    return ws


wb = Workbook()

# ── 00 README ────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "00_Readme"
ws.sheet_properties.tabColor = GOLD
ws["A1"] = "HELSINKI · 16–20 AUG 2026 · FIELD PACK"
ws["A1"].font = title_font
ws.merge_cells("A1:D1")
ws["A2"] = (
    f"Arrive Silja Symphony, Olympia Terminal ~10:30 Sun 16 Aug. "
    f"Leave Thu 20 Aug 15:00 (YOUR ticket decides airport vs ferry — this file does not). "
    f"Research accessed {ACCESSED}. Hours and prices MOVE. Green = bookable tonight. "
    f"Night of the Arts is Thu 20 evening: you leave 15:00 so you miss it. "
    f"Amos Rex closed Tue 18. Grön closed Sun–Tue. Nolla closed Sun–Mon. "
    f"Latitude 25 closed Sun–Tue. Momotoko ramen is dead (bankruptcy 2025). "
    f"Thesis: fortress island · underground museum · sauna on the sea · Georgian table · one Nordic kitchen if a table exists."
)
ws["A2"].alignment = wrap
ws.merge_cells("A2:D5")
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 28
ws.row_dimensions[4].height = 28
ws.row_dimensions[5].height = 28

ws["A7"] = "Sheet"
ws["B7"] = "What's on it"
ws["C7"] = "Do tonight?"
for col in range(1, 4):
    ws.cell(7, col).fill = header_fill
    ws.cell(7, col).font = white_font
sheets_index = [
    ("01_Itinerary", "Hour-level Sun–Thu with meals + backup", "Read once"),
    ("02_Book_tonight", "Ranked bookers with official URLs", "YES — in order"),
    ("03_Restaurants", "Neighbourhood + high-end + ramen + sushi + skip", "Filter Book=Yes"),
    ("04_Arrival_brunch", "First meal after the ship — Ekberg is #1", "Book Ekberg 11:30"),
    ("05_Activities", "Museums, island, sauna, civic", "Rex Sun / island Mon"),
    ("06_Tours_bookable", "Guided tours that actually exist this week", "Optional island tour"),
    ("07_Open_matrix", "Who is open Sun–Thu", "Don't fight closures"),
    ("08_Rain_and_skip", "Weather swaps and tourist traps", "—"),
    ("09_Viz_data", "Must-scores + chart", "—"),
    ("10_Sources", "Official URLs, accessed 15 Aug 2026", "Re-verify volatile cells"),
]
for i, row in enumerate(sheets_index, 8):
    for j, v in enumerate(row, 1):
        ws.cell(i, j, v).font = body
        ws.cell(i, j).fill = cream_fill if i % 2 else pale_fill
        ws.cell(i, j).alignment = wrap
        ws.cell(i, j).border = thin

ws["A19"] = "Locked days (do not rearrange)"
ws["A19"].font = section_font
ws.merge_cells("A20:D24")
ws["A20"] = (
    "SUN 16  Ekberg/Fazer brunch → Amos Rex (closed Tue; Sun 11–17) → Oodi (Sun 10–20) → Georgian Kitchen.\n"
    "MON 17  Suomenlinna (HSL ferry, Blue Route). Evening: Katana ramen (likely closed Sun) or Kallio.\n"
    "TUE 18  ADM Design Museum (Rex closed). Lunch: Levain Ullanlinna next door. Dinner: Nolla.\n"
    "WED 19  Löyly sauna + Baltic. Dinner: Grön if the table exists, else Latitude 25 / Shii.\n"
    "THU 20  Buffer. Ateneum is FREE 10:00–20:00 this day — only if departure mode allows. Out by 15:00."
)
ws["A20"].alignment = wrap
ws.row_dimensions[20].height = 90
autosize(ws, [22, 62, 28, 20])

# ── 01 ITINERARY ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("01_Itinerary")
itin_h = ["Day", "Date", "Block", "Time", "Plan", "Neighbourhood", "Eat", "Book?", "Backup", "Notes"]
itin = [
    ["1 Sun", "2026-08-16", "Arrive", "10:30–11:15",
     "Olympia Terminal → walk/tram into Punavuori. Bags if hotel ready; else cloak VERIFY hotel. Do not museum yet.",
     "South Harbour", "Fruit from the ship is a snack, not lunch.", "Walk",
     "Any bakery on Bulevardi if Ekberg queue", "Inside cabin hangover. Terminal food = bad first impression."],
    ["1 Sun", "2026-08-16", "Brunch", "11:30–13:30",
     "EKBERG CAFÉ, Bulevardi 9. Oldest café (1852). Official summer café Sun 9–18 (9.8–6.9.2026). Official brunch Sat–Sun 9:00–14:30. Book the table.",
     "Punavuori", "Brunch plate / coffee + savoury. VERIFY buffet still running.",
     "BOOK ekberg.fi table reservation (Quandoo)",
     "Fazer Kluuvikatu 3 (Sun 10–20) or Levain Merikortteli Sat–Sun 8:30–16:30",
     "This is the first impression. ~15 min walk from Olympia via Design District."],
    ["1 Sun", "2026-08-16", "Museum", "13:45–16:45",
     "AMOS REX, Mannerheimintie 22–24. Closed Tuesday so do it TODAY. Sun 11–17 — do not linger past 16:30. Generation 2026 (13 May–6 Sep 2026).",
     "Kamppi / Lasipalatsi", "Museum café only if still hungry",
     "Ticket amosrex.fi — 18–29 / students €5 official",
     "Oodi if Rex queue insane",
     "Night of the Arts late opening is Thu 20 11–23 — you will have left."],
    ["1 Sun", "2026-08-16", "Civic", "17:00–18:30",
     "Oodi, Töölönlahdenkatu 4. Free. Official Sun 10–20. Then shower/reset.",
     "Kansalaistori", "—", "No", "Walk Töölönlahti", "Don't make a day of Temppeliaukio."],
    ["1 Sun", "2026-08-16", "Dinner", "19:00–21:30",
     "GEORGIAN KITCHEN, Albertinkatu 7. Family. Official Sun 13–22. Prefer over Rioni (chain, also Sun 13–22, Kasarmikatu 25).",
     "Punavuori", "Khachapuri + khinkali + amber wine. Don't cut khinkali.",
     "BOOK +358 50 383 3228 / georgiankitchen.fi",
     "Rioni; Gaijin Bulevardi 6 from 15:00 (kitchen Sun to 21:00)",
     "Gaijin is better cooking, not Georgian. Lunch there is closed 22.6–17.8.2026."],
    ["2 Mon", "2026-08-17", "Island", "09:30–15:30",
     "SUOMENLINNA. HSL ferry from east side of Kauppatori (by Presidential Palace), ~15 min, route 19. AB ticket covers it (adult single €3.30 on hsl.fi 15 Aug — VERIFY). Blue Route to King's Gate. Picnic. Not Vallisaari.",
     "Suomenlinna", "Island café / picnic. Not a destination restaurant day.",
     "HSL app BEFORE the pier. Optional English tour 10:30/12:30/14:30 €16",
     "If storm: swap with Tue museums",
     "No fortress entrance fee. Last ferries run late; still don't cut dinner."],
    ["2 Mon", "2026-08-17", "Evening", "17:30–21:00",
     "KATANA RAMEN, Telakkakatu 12 — dedicated ramen after Momotoko died. Falstaff lists Mon–Sat 11–21, Sunday omitted → VERIFY, likely closed Sun so TODAY is the ramen day. Or Kallio walk + Nepalese.",
     "Punavuori / Telakka or Kallio", "Tonkotsu at Katana; or dal/momo in Kallio",
     "Walk-in likely. Georgian Supra closed Monday (TableOnline — VERIFY).",
     "Bui Ramen is real but Kalasatama (Aallonhalkoja 1). Ichiraku Kamppi 5F = emergency mall bowl.",
     "Momotoko is CLOSED. Do not go to Yliopistonkatu 5."],
    ["3 Tue", "2026-08-18", "Design", "11:00–16:00",
     "ARCHITECTURE & DESIGN MUSEUM, Korkeavuorenkatu 23. Weekdays 11–20. Square construction from 3 Aug 2026; doors still open. Then Design District drift (shops, not souvenirs).",
     "Kaartinkaupunki / Punavuori",
     "Lunch: Levain Ullanlinna, Korkeavuorenkatu 21, daily 8–16. Or Gaijin if lunch has resumed after 17 Aug — VERIFY.",
     "Museum ticket admuseo.fi (18–29 €6 official)",
     "Ateneum if you want Nationalmuseum analogue (Tue 10–18)",
     "AMOS REX CLOSED TODAY. Do not show up."],
    ["3 Tue", "2026-08-18", "Dinner", "17:30–22:00",
     "NOLLA, Fredrikinkatu 22. Zero-waste Nordic, Bib Gourmand. Closed Sun–Mon so TUE is the first night it opens for you. Confirm hours + price in the booker (official site does not print a hours table).",
     "Punavuori", "Chef's Choice / seasonal",
     "BOOK https://restaurantnolla.com/reservations/",
     "Sea Horse Kapteeninkatu 11 daily 15–24; Kosmos Kalevankatu 3 Mon–Fri 11:30–24",
     "This is the AG-equivalent 'better cooking' without blowing €188."],
    ["4 Wed", "2026-08-19", "Sauna", "13:00–16:00",
     "LÖYLY, Hernesaarenranta 4. Official sauna Wed 9–11 & 13–22. €29 / 2 h inc towel (official). Swimwear required in mixed areas (€8 rental). Swim weather-dependent. Book a 2-hour slot.",
     "Hernesaari", "Löyly kitchen Mon–Sat 11–23 — salmon soup after, not a second fine dining",
     "BOOK varaus.asio.fi Löyly calendar",
     "Allas Sea Pool, Katajanokanlaituri 2A, Mon–Thu 6:30–21",
     "24h cancel. Flip-flops help. Architecture you can sit in."],
    ["4 Wed", "2026-08-19", "Dinner", "17:00–22:00",
     "GRÖN, Albertinkatu 36 — YOUR ONLY NIGHT (closed Sun–Tue). Official summer 2026 menu 188e (omnivore or vegan preorder). Else Latitude 25 omakase (Albertinkatu 19, Wed–Thu 17–21:30) or Shii (Fabianinkatu 17, 16 seats, not vegetarian).",
     "Punavuori", "Grön set 188e / Latitude omakase / Shii 11-course",
     "BOOK DinnerBooking Grön; latitude25.fi; shii.fi — take whichever confirms",
     "Nolla if you saved it; Gaijin; Sea Horse",
     "Olo/Palace/Finnjävel blow more. Finnjävel Tue–Sat; skip unless you decide to."],
    ["5 Thu", "2026-08-20", "Buffer", "09:00–12:30",
     "Easy. Ateneum is FREE today 10:00–20:00 (official free-admission day) — only if YOUR departure mode allows. Or Oodi reprise. Do NOT start Suomenlinna. Night of the Arts starts after you leave.",
     "Centre", "Coffee + pulla. No heavy lunch if flying.",
     "No paid booking needed for Ateneum today",
     "If 15:00 is a FERRY: be at terminal with YOUR ticket's boarding rule. If AIRPORT: leave the centre with YOUR transfer time.",
     "Amos Rex 11–23 tonight — you miss it. ADM architecture walk 15:00 today — collision, do not book."],
    ["5 Thu", "2026-08-20", "Leave", "by 15:00",
     "Out. Confirm mode on YOUR ticket.",
     "—", "Airport / terminal food only as emergency", "—", "—",
     "Do not cut boarding like the Stockholm kiosk lesson."],
]
add_table(ws, itin_h, itin, [10, 12, 12, 14, 58, 22, 40, 36, 36, 40], CORAL, 78)
for r in range(2, ws.max_row + 1):
    if "Dinner" in str(ws.cell(r, 3).value) or "Brunch" in str(ws.cell(r, 3).value):
        ws.cell(r, 7).fill = must_fill
    if "BOOK" in str(ws.cell(r, 8).value).upper():
        ws.cell(r, 8).fill = book_fill

# ── 02 BOOK TONIGHT ──────────────────────────────────────────────────────────
ws = wb.create_sheet("02_Book_tonight")
book_h = ["Priority", "What", "When", "How", "URL", "Phone", "If sold out", "Must?"]
books = [
    ["0", "Ekberg brunch", "Sun 16 ~11:30", "Quandoo via official page; 1.5 h; max 6",
     "https://www.ekberg.fi/en/cafe/table-reservation", "+358 9 6811 860",
     "Walk-in café / Fazer / Levain Merikortteli", "YES — first impression"],
    ["1", "Georgian Kitchen", "Sun 16 ~19:00", "Site form / phone",
     "https://www.georgiankitchen.fi/", "+358 50 383 3228",
     "Rioni +358 50 551 2264 · https://www.rioni.fi/en-gb/helsinki", "YES — the cuisine you named"],
    ["2", "Nolla", "Tue 18 dinner", "Official reservations page",
     "https://restaurantnolla.com/reservations/", "+358 40 163 9313 (Michelin listing)",
     "Sea Horse / Kosmos", "YES — best cooking under star-tier"],
    ["3", "Grön", "Wed 19 ONLY", "DinnerBooking",
     "https://dinnerbooking.com/fi/en-US/r3904/restaurant-gron", "VERIFY on booker",
     "Latitude 25 / Shii / unused Nolla", "If you want the 188e summer menu"],
    ["4", "Löyly 2h public sauna", "Wed 19 13:00+", "Official calendar",
     "https://varaus.asio.fi/onlinekalenteri/loyly/guest.php?ss_lang=eng", "+358 50 4768741 sauna",
     "Allas Sea Pool", "YES — architecture + Baltic"],
    ["5", "Gaijin", "Sun from 15:00 or later", "Official reservation",
     "https://www.gaijin.fi/reservation", "010 322 9386 / same-day 010 322 9381",
     "Nolla", "Strong alt, not Georgian"],
    ["6", "Latitude 25", "Wed 19 (closed Sun–Tue)", "Call / site",
     "https://www.latitude25.fi/", "+358 9 6128 6000",
     "Shii https://shii.fi/ 050 329 0360", "Only if Grön fails and you want omakase"],
    ["7", "Sea Horse", "Any dinner from 15:00", "Online / phone",
     "https://www.seahorse.fi/en/restaurant-sea-horse/", "+358 9 628 169",
     "Walk-in off-peak possible", "Finnish classic backup"],
    ["8", "Amos Rex ticket", "Sun 16 before 17:00", "Official tickets",
     "https://amosrex.fi/en/tickets/", "09 6844 460",
     "Door; 18–29 / students €5 official", "YES today — closed Tue"],
    ["9", "HSL ticket / app", "All week inc Suomenlinna ferry", "HSL app or pier machine BEFORE boarding",
     "https://www.hsl.fi/en/travelling/visitors/suomenlinna", "—",
     "Contactless at readers", "YES"],
    ["10", "Optional: Suomenlinna English tour", "Mon 17 10:30 / 12:30 / 14:30", "Ehrensvärd Society shop",
     "https://suomenlinna.johku.com/en_US/liput/guided-walking-tour-in-english", "+358 9 68999 850",
     "Self-guided Blue Route is enough", "Optional"],
    ["11", "Do NOT book", "Cathedral tours, SkyWheel, food tours, Temppeliaukio as a day, ADM walk Thu 15:00",
     "—", "—", "—", "—", "Skip"],
]
add_table(ws, book_h, books, [10, 28, 24, 36, 55, 28, 36, 28], "2E7D32", 44)
for r in range(2, ws.max_row + 1):
    url = ws.cell(r, 5).value
    if url and str(url).startswith("http"):
        href(ws, r, 5, url)
    must = str(ws.cell(r, 8).value or "")
    if must == "Skip":
        for c in range(1, 9):
            ws.cell(r, c).fill = skip_fill
    elif r <= 6:
        ws.cell(r, 1).fill = book_fill

# ── 03 RESTAURANTS ───────────────────────────────────────────────────────────
ws = wb.create_sheet("03_Restaurants")
rh = ["Name", "Cuisine", "Neighbourhood", "Address", "Tier", "Price", "Book",
      "Sun16", "Mon17", "Tue18", "Wed19", "Thu20", "Why Bob", "Order / note", "URL", "Hours source", "Must"]
rests = [
    ["Ekberg Café", "Bakery / brunch", "Punavuori", "Bulevardi 9", "icon", "VERIFY", "Book brunch",
     "Brunch 9–14:30; café 9–18", "Breakfast 7:30–10:30; café 7:30–19", "same", "same", "same",
     "Oldest café — first impression without tourist log-cabin", "Brunch plate",
     "https://www.ekberg.fi/", "Official 9.8–6.9.2026 summer hours", 10],
    ["Karl Fazer Café", "Café / chocolate", "Kluuvi", "Kluuvikatu 3", "icon", "VERIFY", "Walk-in",
     "10–20", "7:30–22", "7:30–22", "7:30–22", "7:30–22",
     "1891 room; weekend brunch advertised, sitting hours not printed", "Salmon soup + pastry VERIFY",
     "https://www.fazer.fi/fazer-cafe/kahvilat/karl-fazer-cafe/", "Official café hours", 8],
    ["Levain Merikortteli", "Bakery / eatery", "Punavuori", "Pursimiehenkatu 29–31", "neighbourhood", "VERIFY", "Walk-in / book",
     "8:30–16:30", "8:00–16:30", "8:00–16:30", "8:00–16:30", "8:00–16:30",
     "Sourdough backup if Ekberg slammed", "Cardamom bun + savoury",
     "https://www.levain.fi/contact", "Official", 6],
    ["Levain Ullanlinna", "Bakery / eatery", "Ullanlinna", "Korkeavuorenkatu 21", "neighbourhood", "VERIFY", "Walk-in",
     "8–16", "8–16", "8–16", "8–16", "8–16",
     "Next door to ADM — Tuesday lunch", "Sourdough lunch",
     "https://www.levain.fi/contact", "Official", 7],
    ["Georgian Kitchen", "Georgian", "Punavuori", "Albertinkatu 7", "neighbourhood", "VERIFY", "Book",
     "13–22", "15–23", "15–23", "15–23", "15–23",
     "Family table; prefer over the chain", "Khachapuri + khinkali",
     "https://www.georgiankitchen.fi/", "Official", 10],
    ["Rioni", "Georgian", "Kaartinkaupunki", "Kasarmikatu 25", "hot", "VERIFY", "Book",
     "13–22", "16–23", "16–23", "16–23", "16–23",
     "Open Sunday; chain; courtyard", "Same supra logic",
     "https://www.rioni.fi/en-gb/helsinki", "Official", 8],
    ["Georgian Supra", "Georgian / wine", "Kallio", "Hämeentie 36", "hot", "VERIFY", "Book",
     "14–21*", "CLOSED*", "16–22*", "16–22*", "16–22*",
     "Kallio night energy. Hours from TableOnline 12.5.2026 — VERIFY", "Supra + qvevri",
     "https://www.georgiansupra.fi/", "Official address; hours secondary", 7],
    ["Gaijin", "North Asia / Levant-leaning sharing", "Punavuori", "Bulevardi 6", "hot", "VERIFY", "Book",
     "Dinner 15:00; kitchen 21:00", "Dinner 16:30 (lunch closed)", "Lunch MAY resume; dinner 16:30", "Dinner 16:30", "Dinner 16:30",
     "Serious cooking next to Ekberg. Not Georgian. Lunch closed 22.6–17.8.2026.", "Sharing plates",
     "https://www.gaijin.fi/reservation", "Official contact/reservation", 8],
    ["Nolla", "Nordic zero-waste", "Punavuori", "Fredrikinkatu 22", "icon", "VERIFY in booker", "Book",
     "CLOSED*", "CLOSED*", "Dinner ~17:30*", "Dinner*", "Leave 15:00",
     "Bib Gourmand; thesis kitchen; official site has no hours table", "Chef's Choice",
     "https://restaurantnolla.com/reservations/", "Hours: Michelin listing (secondary)", 10],
    ["Grön", "Forage Nordic", "Punavuori", "Albertinkatu 36", "icon", "188e menu official", "Book (hard)",
     "CLOSED", "CLOSED", "CLOSED", "Dinner*", "Leave 15:00",
     "Official summer 2026 menu 188e + vegan 188e preorder. Wine pairing 142e.", "Set (request vegan in booker if needed)",
     "https://dinnerbooking.com/fi/en-US/r3904/restaurant-gron", "Menu official; hours secondary", 9],
    ["Latitude 25", "Omakase", "Punavuori", "Albertinkatu 19", "icon", "VERIFY", "Book",
     "CLOSED", "CLOSED", "CLOSED", "17:00–21:30", "Leave 15:00",
     "Billed as Finland's premier omakase. Only Wednesday for you.", "Omakase counter",
     "https://www.latitude25.fi/", "Official hours", 7],
    ["Shii", "Omakase / otsumami", "Kaartinkaupunki", "Fabianinkatu 17", "icon", "VERIFY", "Book",
     "VERIFY", "VERIFY", "VERIFY", "VERIFY", "Leave 15:00",
     "16 seats; 11-course; cannot be made vegetarian", "Counter",
     "https://shii.fi/", "Official concept; hours missing — VERIFY", 7],
    ["Domo", "Sushi à la carte", "Kamppi", "Kalevankatu 21", "neighbourhood", "VERIFY", "Walk-in / call",
     "VERIFY", "VERIFY", "VERIFY", "VERIFY", "VERIFY",
     "Long-running; more fish less rice; not buffet", "Nigiri",
     "VERIFY official site", "Hours not captured — VERIFY", 6],
    ["Katana Ramen", "Ramen", "Punavuori / Telakka", "Telakkakatu 12", "hot", "VERIFY", "Walk-in",
     "CLOSED?*", "11–21*", "11–21*", "11–21*", "11–21*",
     "Dedicated ramen after Momotoko died. HS reviewed Jul 2026.", "Tonkotsu",
     "katanaramenclub@gmail.com", "Falstaff hours (secondary)", 8],
    ["Bui Ramen", "Ramen / mazemen", "Kalasatama", "Aallonhalkoja 1 L T 2", "neighbourhood", "Bui Paitan 17.90€ official", "Walk-in / book",
     "12–20", "11–21", "11–21", "11–21", "11–21",
     "Handmade noodles; not central — only if east", "Bui Paitan",
     "https://buiramen.fi/", "Official", 5],
    ["Ramen Ichiraku", "Ramen", "Kamppi mall 5F", "Urho Kekkosen katu 1", "cheap-everyday", "VERIFY", "Walk-in",
     "12–19", "11–21", "11–21", "11–21", "11–21 morning",
     "Emergency bowl. Mall. Time Out: reliable, not soulful.", "Tonkotsu / tan tan",
     "https://www.kamppihelsinki.fi/en/shops-and-services/other-stores/ramen-ichiraku/-/110", "Official Kamppi listing", 5],
    ["Sushi Wagocoro", "Sushi neighbourhood", "Taka-Töölö", "Runeberginkatu 63 A 21", "neighbourhood", "VERIFY", "Book/call — tiny",
     "VERIFY", "VERIFY", "VERIFY", "VERIFY", "VERIFY",
     "Japanese-owned small room. Confirm before crossing town.", "Nigiri",
     "https://www.myhelsinki.fi/places/sushi-wagocoro/", "MyHelsinki — hours VERIFY", 6],
    ["Kabuki", "Traditional Japanese", "Kamppi", "Lapinlahdenkatu 12", "neighbourhood", "VERIFY", "Book — small room",
     "VERIFY", "VERIFY", "VERIFY", "VERIFY", "VERIFY",
     "Old-school Japanese / tatami, not Nordic omakase theatre", "Sushi / set",
     "https://www.myhelsinki.fi/places/kabuki/", "MyHelsinki — hours VERIFY", 6],
    ["Sansar (ex-Satkar)", "Nepalese", "Kamppi", "Fredrikinkatu 46", "cheap-everyday", "VERIFY", "Walk-in lunch",
     "12–21", "10:30–21; lunch to 16", "same", "same", "lunch if time",
     "Kamppi since 1998. Ordinary diaspora lunch. Search Sansar, not Satkar.", "Tandoor / curry lunch",
     "https://www.sansar.fi/", "Official", 7],
    ["Restaurant Mountain", "Nepalese", "Taka-Töölö", "Nordenskiöldinkatu 8", "cheap-everyday", "VERIFY", "Walk-in lunch",
     "12:30–22", "10:30–22:30; lunch to 15", "same", "same", "lunch if time",
     "Family-owned since 2000. The kebabpizza analogue.", "Dal / tandoor",
     "https://ravintolamountain.com/en/home/", "Official", 7],
    ["BasBas (Baskeri & Basso)", "Bistro / natural wine", "Punavuori", "Tehtaankatu 27–29 courtyard", "hot", "VERIFY", "Book",
     "CLOSED", "CLOSED", "from 16:00", "from 16:00", "too late",
     "Ingredient-led small plates. Closed Sun–Mon so Tue/Wed.", "Share plates + wine",
     "https://basbas.fi/bistro/en/", "Official", 7],
    ["Nolita", "Seasonal bistro / wine", "Punavuori", "Kankurinkatu 6", "hot", "VERIFY", "Book dinner",
     "CLOSED", "CLOSED", "17–22", "17–22", "too late",
     "Closed Sun–Mon. In-house sourdough with dinner.", "Shareable bistro",
     "https://nolita.fi/en/", "Official", 6],
    ["The Alley", "Small plates / wine", "Kallio", "Fleminginkatu 8", "hot", "VERIFY", "Book / some walk-in",
     "CLOSED", "CLOSED", "17–23 kitchen 21:30", "17–23 kitchen 21:30", "too late",
     "Kallio-scale BasBas cousin. Closed Sun–Mon.", "Small plates",
     "https://www.thealleyhelsinki.com/", "Official", 6],
    ["Way Bakery Kallio", "Bakery / wine", "Kallio", "Agricolankatu 9", "neighbourhood", "VERIFY", "No reservations",
     "09–16", "08–19", "08–19", "08–19", "08–19 morning",
     "Sourdough that can hold a Kallio morning. Not a Fazer clone.", "Bread + coffee; wine later",
     "https://www.waybakery.fi/kallio", "Official", 7],
    ["Café Succès", "Traditional café", "Ullanlinna", "Korkeavuorenkatu 2", "neighbourhood", "VERIFY", "Walk-in",
     "11–17", "08–18", "08–18", "08–18", "08–18 morning",
     "1957 family café; giant cinnamon bun. Near ADM.", "Korvapuusti",
     "https://succes.fi/en/frontpage/", "Official", 6],
    ["Eromanga", "Bakery / lihapiirakka", "Kaartinkaupunki", "Pohjoinen Makasiinikatu 6", "neighbourhood", "VERIFY", "Walk-in",
     "CLOSED", "07:30–15:30", "07:30–15:30", "07:30–15:30", "07:30–15:30 if buffer",
     "1946-recipe meat pie. Savoury bakery, not another bun.", "Lihapiirakka",
     "https://eromanga.fi/home/", "Official", 5],
    ["Kannas", "Finnish port classic", "Hietalahti", "Eerikinkatu 43", "neighbourhood", "VERIFY", "Book dinner",
     "14–24 kitchen 22:30", "16–24 kitchen 22:30", "same", "same", "too late",
     "Since 1939; sailors/dockworkers room. Less polished than Sea Horse.", "Finnish classics",
     "https://kannas.fi/english/", "Official", 6],
    ["Sea Horse", "Classic Finnish", "Ullanlinna", "Kapteeninkatu 11", "icon", "VERIFY", "Book / walk-in",
     "15–24", "15–24", "15–24", "15–24", "15–24 (too late to start)",
     "1933 institution; onion steak; near ADM", "Onion steak / Baltic herring / vorschmack",
     "https://www.seahorse.fi/en/restaurant-sea-horse/", "Official summer 8.6.2026+", 8],
    ["Kosmos", "Helsinki classic", "Kamppi", "Kalevankatu 3", "icon", "Menu printed; VERIFY", "Phone / email",
     "CLOSED", "11:30–24", "11:30–24", "11:30–24", "11:30–24",
     "1924; blinis, vorschmack, Baltic herring. Closed Sundays.", "Fried Baltic herrings 22€ on fetched menu",
     "https://kosmos.fi/en/our-food/", "Official", 7],
    ["Elite", "Artist restaurant", "Töölö", "Eteläinen Hesperiankatu 22", "icon", "VERIFY", "Book",
     "13–22", "12–22", "12–22", "12–23", "12–23",
     "Since 1932; onion steak cousin in Töölö", "Classic Finnish",
     "https://elite.fi/en/front-page/", "Official", 6],
    ["Yes Yes Yes", "Vegetarian seasonal", "Punavuori", "Iso Roobertinkatu 1", "hot", "VERIFY", "Walk-in / book",
     "VERIFY", "VERIFY", "VERIFY", "VERIFY", "VERIFY",
     "Local-farm veg; wine. Hours not on official info page.", "Seasonal plates",
     "https://yesyesyes.fi/", "Official address; hours VERIFY", 6],
    ["Löyly Restaurant", "Finnish casual / terrace", "Hernesaari", "Hernesaarenranta 4", "neighbourhood", "VERIFY", "Walk-in terrace",
     "Kitchen 11–22", "Kitchen 11–23", "Kitchen 11–23", "Kitchen 11–23", "Kitchen 11–23",
     "After sauna — not a destination kitchen", "Lohikeitto",
     "https://www.loylyhelsinki.fi/en/info", "Official kitchen hours", 6],
    ["Finnjävel Salonki", "Finnish fine dining", "Kamppi", "Ainonkatu 3", "icon", "star-tier VERIFY", "Book",
     "CLOSED", "CLOSED", "17–23", "17–23", "Leave 15:00",
     "Michelin; skip unless you decide to blow the budget. Summer holiday already over.", "Tasting",
     "https://finnjavel.fi/en/", "Official hours", 3],
    ["Olo / Palace", "Nordic 1–2★", "Harbour", "VERIFY live", "icon", "200+ VERIFY", "Book weeks ahead",
     "VERIFY", "VERIFY", "VERIFY", "VERIFY", "Leave 15:00",
     "Blow-budget. Skip unless you choose to.", "Tasting",
     "Michelin / official sites", "Not fetched in full this pass", 2],
    ["Nepalese (Kallio)", "Nepalese", "Kallio", "Pick a busy room", "cheap-everyday", "VERIFY", "Walk-in",
     "Many open", "Many open", "Many open", "Many open", "Light only",
     "Helsinki's kebabpizza analogue — ordinary diaspora lunch", "Dal bhat / momo",
     "Walk Hämeentie", "Neighbourhood observation, not a single venue", 6],
    ["Savotta / Zetor / Kappeli", "Tourist Finnish", "Senate Square / Esplanadi", "—", "skip", "—", "—",
     "—", "—", "—", "—", "—",
     "Postcard Finland", "Don't", "—", "—", 0],
    ["Sushi buffet chains", "Buffet", "Everywhere", "—", "skip", "—", "—",
     "—", "—", "—", "—", "—",
     "Not Domo", "Don't", "—", "—", 0],
    ["Momotoko", "Ramen CLOSED", "Yliopistonkatu 5 (former)", "—", "skip", "—", "—",
     "DEAD", "DEAD", "DEAD", "DEAD", "DEAD",
     "Cathywong Oy bankruptcy May 2025; chain said permanently closed", "Do not go",
     "https://www.hs.fi/helsinki/art-2000011230551.html", "HS 15.5.2025 + MTV 31.5.2025", 0],
    ["Wino", "Wine bar CLOSED this trip", "Fleminginkatu 11", "—", "skip", "—", "—",
     "CLOSED 1.6–3.9.2026", "same", "same", "same", "same",
     "Facade renovation. Do not go.", "Don't",
     "https://www.wino.fi/restaurant-wino/", "Official closure notice", 0],
    ["Restaurant Kuu", "Töölö classic CLOSED", "Töölönkatu 27", "—", "skip", "—", "—",
     "CLOSED from 1.7.2026", "same", "same", "same", "same",
     "Housing-company renovation; reopen autumn. Don't google-walk there.", "Don't",
     "https://ravintolakuu.fi/en/kuu-closed-from-1-7/", "Official", 0],
    ["Adzika / Satkar / Sandro", "Moved or dead names", "stale maps", "—", "skip", "—", "—",
     "TRAP", "TRAP", "TRAP", "TRAP", "TRAP",
     "Adzika seeking premises; Satkar is now Sansar; Sandro closed 28 Feb.", "Search current names",
     "See research/neighbourhood-eateries.md", "Official closure pages", 0],
]
add_table(ws, rh, rests, [20, 22, 18, 24, 14, 16, 16, 22, 18, 20, 18, 18, 40, 28, 40, 28, 8], TEAL, 52)
for r in range(2, ws.max_row + 1):
    url = ws.cell(r, 15).value
    if url and str(url).startswith("http"):
        href(ws, r, 15, url)
    tier = ws.cell(r, 5).value
    if tier == "skip":
        for c in range(1, 18):
            ws.cell(r, c).fill = skip_fill
    elif tier == "icon":
        ws.cell(r, 5).fill = gold_fill
    if str(ws.cell(r, 7).value or "").startswith("Book"):
        ws.cell(r, 7).fill = book_fill
    rank = ws.cell(r, 17).value
    if isinstance(rank, int) and rank >= 9:
        ws.cell(r, 1).fill = must_fill

# ── 04 BRUNCH ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("04_Arrival_brunch")
bh = ["Rank", "Place", "Walk from Olympia", "Sunday published", "Why first meal", "Book?", "Skip if"]
brunch = [
    ["1", "Ekberg Café, Bulevardi 9", "~12–18 min walk via Design District",
     "Café 9–18 (9.8–6.9.2026); BRUNCH 9:00–14:30",
     "Oldest bakery-café; you walk into the city, not a terminal. Official table reservation.",
     "YES — ekberg.fi table reservation", "You need waterfront immediately (you don't)"],
    ["2", "Karl Fazer Café, Kluuvikatu 3", "~20 min via Esplanadi / Kauppatori",
     "Café 10–20; weekend brunch advertised, sitting hours VERIFY",
     "1891 chocolate-café institution", "Walk-in; table link on official page",
     "Queue out the door — then Ekberg"],
    ["3", "Levain Merikortteli, Pursimiehenkatu 29–31", "Punavuori, slightly further than Ekberg",
     "Sat–Sun 8:30–16:30 official", "Third-wave sourdough if institutions are slammed", "Walk-in / book",
     "You want a tablecloth first impression"],
    ["4", "Gaijin, Bulevardi 6", "Next to Ekberg — NOT brunch",
     "Dinner from 15:00 Sun; lunch closed through 17.8.2026",
     "Park for 15:00 if Ekberg was too light", "Book gaijin.fi",
     "You insist on eating at 11:30"],
    ["5", "Café Regatta, Töölö shore", "TOO FAR (~40 min) for arrival", "VERIFY",
     "Cute; save for a Töölö walk another day", "No", "You are tired from the ship"],
    ["6", "Terminal / leftover fruit", "0 min", "Always", "Emergency only", "No",
     "You want Helsinki to begin well"],
]
add_table(ws, bh, brunch, [8, 36, 38, 40, 50, 32, 28], GOLD, 44)
for c in range(1, 8):
    ws.cell(2, c).fill = book_fill
href(ws, 2, 6, "https://www.ekberg.fi/en/cafe/table-reservation", "YES — ekberg.fi table reservation")

# ── 05 ACTIVITIES ────────────────────────────────────────────────────────────
ws = wb.create_sheet("05_Activities")
ah = ["Name", "Type", "Time", "Book?", "Hours / catch this week", "URL", "Fit vs Stockholm", "Must", "Price if official"]
acts = [
    ["Amos Rex", "Museum", "2–3 h", "Ticket",
     "Closed Tue. Sun 11–17. Night of Arts 20 Aug 11–23 (you miss). Generation 2026: 13 May–6 Sep 2026.",
     "https://amosrex.fi/en/visit-us/", "Fotografiska / Moderna analogue", "YES Sun",
     "Door €22 / online €20; 18–29 & students €5"],
    ["Architecture & Design Museum", "Museum", "2–3 h", "Ticket",
     "Mon–Fri 11–20; Sat–Sun 11–18. Square construction from 3 Aug 2026; doors open. Night of Arts 20 Aug 11–22 (you miss).",
     "https://admuseo.fi/en/visit-us/", "ArkDes analogue", "YES Tue",
     "Adults €23; 18–29 €6"],
    ["Oodi", "Library / architecture", "45–90 min", "Free",
     "Mon–Fri 8–21; Sat–Sun 10–20. No 16–20 Aug exception on the 2026 list.",
     "https://oodihelsinki.fi/en/arrival/", "Civic design", "YES", "Free"],
    ["Suomenlinna", "UNESCO fortress island", "4–5 h", "HSL ferry",
     "Ferry Kauppatori east side ~15 min, route 19. AB ticket. Blue Route to King's Gate. No entrance fee.",
     "https://suomenlinna.fi/en/explore/arriving/", "Vasa-scale object that is also a neighbourhood", "YES Mon",
     "HSL AB single €3.30 on 15 Aug page — VERIFY"],
    ["Löyly public sauna", "Sauna + architecture", "2 h slot", "Book",
     "Wed 9–11 & 13–22. Swimwear mixed. Swim weather-dependent. Hernesaarenranta 4.",
     "https://www.loylyhelsinki.fi/en/public-sauna", "A building you can sit in", "YES Wed",
     "€29 / 2 h inc towel; extra hour €12; swimwear rental €8"],
    ["Ateneum", "Art museum", "2 h", "Ticket / free Thu",
     "Jul–Aug open daily. Sun 10–17; Mon 11–17; Tue 10–18; Wed–Thu 10–20. FREE Thu 20 Aug 10–20.",
     "https://ateneum.fi/en/opening-hours-and-tickets/", "Nationalmuseum analogue", "Optional Thu morning",
     "Door €23 / online €21; free 20 Aug"],
    ["Allas Sea Pool", "Swim / sauna backup", "up to 3 h ticket", "Ticket",
     "Mon–Thu 6:30–21. Katajanokanlaituri 2A. More central, less architecture.",
     "https://allasseapool.fi/en/", "Backup sauna", "Backup",
     "Weekend single €17 official; weekday VERIFY"],
    ["Temppeliaukio", "Church", "20 min", "Door ticket",
     "Hours change with services; check the week. Sunday 10:00 is worship, not sightseeing.",
     "https://www.temppeliaukiochurch.fi/en/index/nimi.html", "Photo only", "No",
     "Adult €8 official"],
    ["Helsinki Cathedral / Senate Square", "Postcard", "15 min pass-through", "No",
     "On the walk from harbour — not a day", "—", "Orientation only", "Pass by", "—"],
    ["Seurasaari", "Open-air museum", "half day", "—", "Skansen analogue — you skipped Skansen", "—",
     "Skip this trip", "No", "—"],
    ["Night of the Arts", "City-wide", "evening 20 Aug", "Many free",
     "Confirmed Thu 20 Aug 2026. You leave 15:00.", "https://helsinkifestival.fi/taiteidenyo/en/instructions-for-event-organisers/",
     "FOMO only", "Miss", "—"],
]
add_table(ws, ah, acts, [28, 22, 14, 16, 55, 42, 32, 18, 36], TEAL, 48)
for r in range(2, ws.max_row + 1):
    url = ws.cell(r, 6).value
    if url and str(url).startswith("http"):
        href(ws, r, 6, url)
    must = str(ws.cell(r, 8).value or "")
    if must.startswith("YES"):
        ws.cell(r, 8).fill = book_fill
    if must in ("No", "Skip", "Miss"):
        ws.cell(r, 8).fill = skip_fill

# ── 06 TOURS ─────────────────────────────────────────────────────────────────
ws = wb.create_sheet("06_Tours_bookable")
th = ["Tour", "When it exists this trip", "Book URL", "Price if published", "Fit", "Book?"]
tours = [
    ["Suomenlinna English walking tour (Ehrensvärd Society)",
     "1.6–31.8.2026 daily 10:30, 12:30, 14:30. 1 h. Starts Suomenlinna Museum lobby.",
     "https://suomenlinna.johku.com/en_US/liput/guided-walking-tour-in-english",
     "Adult €16; student/senior €11; child €6. Summer ticket includes Ehrensvärd Museum.",
     "Optional. Blue Route is enough if you read buildings yourself.", "Optional Mon"],
    ["Suomenlinna self-guided Blue Route",
     "Always. Clock tower at Jetty Barracks → church → Great Courtyard → dry dock → tunnels → King's Gate. ~1.5 km one way.",
     "https://suomenlinna.fi/en/sights/",
     "Free (ferry ticket only)",
     "Default.", "Do this"],
    ["University of Helsinki City Centre Campus tour (English)",
     "MyHelsinki lists Mon 17 Aug 17:00–18:00, free, no registration. VERIFY the listing the morning of.",
     "https://www.myhelsinki.fi/events/guided-art-historical-tours-city-centre-campus/",
     "Free",
     "Bonus after the island if you still have legs.", "Optional"],
    ["ADM Historic Helsinki architecture walk",
     "Thu 20 Aug 15:00–16:30. Starts City Hall.",
     "https://admuseo.fi/en/event/historic-helsinki-a-capital-carved-in-stone/",
     "€15",
     "Direct collision with 15:00 departure.", "DO NOT BOOK"],
    ["ADM private architecture walk",
     "On request. €200/group Mon–Sat, €300 Sunday (museum page).",
     "https://admuseo.fi/en/services-for-groups/",
     "€200–300 / group",
     "Not good value next to a normal ADM visit.", "Skip"],
    ["Secret Food Tours Helsinki",
     "Typically Mon–Sat afternoon. Kallio. 3–3.5 h.",
     "https://www.secretfoodtours.com/helsinki/food-tours-helsinki/",
     "Listed €94.99 — VERIFY checkout",
     "Overlaps your restaurant sheet.", "SKIP"],
    ["Heather's Helsinki Fork in Hand",
     "4.5–5 h food walk. Calendar VERIFY.",
     "https://heathershelsinki.com/",
     "Listed €130 — VERIFY",
     "Same overlap.", "SKIP"],
    ["Löyly public sauna (not a tour, but bookable experience)",
     "Wed 19. Two-hour slots.",
     "https://varaus.asio.fi/onlinekalenteri/loyly/guest.php?ss_lang=eng",
     "€29 / 2 h",
     "The Helsinki verb.", "YES"],
]
add_table(ws, th, tours, [42, 50, 55, 40, 40, 16], "6A1B9A", 52)
for r in range(2, ws.max_row + 1):
    url = ws.cell(r, 3).value
    if url and str(url).startswith("http"):
        href(ws, r, 3, url)
    decision = str(ws.cell(r, 6).value or "")
    if "SKIP" in decision.upper() or "NOT" in decision:
        ws.cell(r, 6).fill = skip_fill
    elif decision.startswith("YES") or decision.startswith("Do this"):
        ws.cell(r, 6).fill = book_fill

# ── 07 OPEN MATRIX ───────────────────────────────────────────────────────────
ws = wb.create_sheet("07_Open_matrix")
mh = ["Place", "Sun 16", "Mon 17", "Tue 18", "Wed 19", "Thu 20 (leave 15:00)", "Flag"]
matrix = [
    ["Ekberg brunch", "OPEN 9–14:30", "breakfast only", "breakfast only", "breakfast only", "breakfast only", "Official"],
    ["Fazer café", "OPEN 10–20", "OPEN", "OPEN", "OPEN", "OPEN morning", "Official"],
    ["Georgian Kitchen", "OPEN 13–22", "OPEN 15–23", "OPEN", "OPEN", "OPEN but leaving", "Official"],
    ["Rioni", "OPEN 13–22", "OPEN 16–23", "OPEN", "OPEN", "OPEN but leaving", "Official"],
    ["Georgian Supra", "OPEN?* 14–21", "CLOSED*", "OPEN?*", "OPEN?*", "OPEN?*", "TableOnline — VERIFY"],
    ["Gaijin", "DINNER 15:00", "DINNER 16:30 (no lunch)", "lunch MAY be back", "DINNER 16:30", "DINNER 16:30", "Official summer note"],
    ["Nolla", "CLOSED*", "CLOSED*", "OPEN dinner*", "OPEN dinner*", "too late", "Michelin hours — VERIFY"],
    ["Grön", "CLOSED", "CLOSED", "CLOSED", "OPEN dinner*", "too late", "Secondary hours; 188e menu official"],
    ["Latitude 25", "CLOSED", "CLOSED", "CLOSED", "OPEN 17–21:30", "too late", "Official"],
    ["Sea Horse", "OPEN 15–24", "OPEN 15–24", "OPEN 15–24", "OPEN 15–24", "OPEN but leaving", "Official summer"],
    ["Kosmos", "CLOSED", "OPEN 11:30–24", "OPEN", "OPEN", "OPEN morning/lunch", "Official"],
    ["Elite", "OPEN 13–22", "OPEN 12–22", "OPEN", "OPEN", "OPEN morning", "Official"],
    ["Katana Ramen", "CLOSED?*", "OPEN 11–21*", "OPEN*", "OPEN*", "OPEN* morning", "Falstaff — VERIFY"],
    ["Bui Ramen", "OPEN 12–20", "OPEN 11–21", "OPEN", "OPEN", "OPEN morning", "Official — not central"],
    ["Amos Rex", "OPEN 11–17", "OPEN 11–20", "CLOSED", "OPEN 11–20", "OPEN 11–23 (you miss evening)", "Official"],
    ["ADM", "OPEN 11–18", "OPEN 11–20", "OPEN 11–20", "OPEN 11–20", "OPEN 11–22 (you miss evening)", "Official"],
    ["Oodi", "OPEN 10–20", "OPEN 8–21", "OPEN 8–21", "OPEN 8–21", "OPEN 8–21 morning", "Official"],
    ["Löyly sauna", "OPEN 11–21", "OPEN 12–22", "OPEN 12–22", "OPEN 9–11 & 13–22", "OPEN 12–22 morning?", "Official"],
    ["Ateneum", "OPEN 10–17", "OPEN 11–17", "OPEN 10–18", "OPEN 10–20", "FREE 10–20", "Official; free day 20 Aug"],
    ["Momotoko", "DEAD", "DEAD", "DEAD", "DEAD", "DEAD", "HS/MTV 2025"],
]
add_table(ws, mh, matrix, [22, 22, 22, 20, 24, 28, 32], NAVY, 28)
for r in range(2, ws.max_row + 1):
    for c in range(2, 7):
        val = str(ws.cell(r, c).value or "").upper()
        if "DEAD" in val or val.startswith("CLOSED"):
            ws.cell(r, c).fill = closed_fill
        elif "MAY" in val or "?*" in str(ws.cell(r, c).value) or "VERIFY" in val:
            ws.cell(r, c).fill = maybe_fill
        elif "OPEN" in val or "DINNER" in val or "FREE" in val or "BRUNCH" in val:
            ws.cell(r, c).fill = open_fill

# ── 08 RAIN ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet("08_Rain_and_skip")
rain = [
    ["Rain on island day", "Swap Suomenlinna with ADM + Ateneum + Punavuori; island on the clear day", "Two islands"],
    ["Löyly full", "Allas; or hotel sauna if any", "Skip sauna entirely — it is the Helsinki verb"],
    ["Nolla + Grön + omakase all full", "Sea Horse + Georgian + Katana. Still a great trip.", "Random Esplanadi terrace"],
    ["Hotel not ready 10:30", "Brunch first with bag; ask hotel left luggage VERIFY", "Sit in the terminal 2 hours"],
    ["Tired like Stockholm Friday", "One museum + one dinner. Cut the walk.", "Temppeliaukio + cathedral + market hall stack"],
    ["Katana closed when you arrive", "Ichiraku Kamppi 5F Sun 12–19 Mon–Thu 11–21 (official mall listing). Bui only if already east.", "Hunting a dead Momotoko / Wino / Kuu"],
    ["Thu 15:00 is a ferry", "Be at the terminal with YOUR boarding rule — do not copy Stockholm 16:30 onto this file", "Starting Suomenlinna Thursday"],
    ["Thu 15:00 is a flight", "Leave the centre with YOUR transfer time. Ateneum free morning only if the buffer is real", "Night of the Arts FOMO"],
]
add_table(ws, ["If", "Then", "Never"], rain, [32, 78, 40], CORAL, 36)

# ── 09 VIZ ───────────────────────────────────────────────────────────────────
ws = wb.create_sheet("09_Viz_data")
ws.append(["Place", "Must_0_10", "Kind"])
viz = [
    ("Ekberg brunch", 10, "Meal"),
    ("Nolla", 10, "Meal"),
    ("Suomenlinna", 10, "Place"),
    ("Amos Rex", 9, "Place"),
    ("Georgian Kitchen", 9, "Meal"),
    ("Löyly", 9, "Place"),
    ("ADM Design Museum", 9, "Place"),
    ("Grön", 9, "Meal"),
    ("Katana Ramen", 8, "Meal"),
    ("Sea Horse", 8, "Meal"),
    ("Gaijin", 8, "Meal"),
    ("Oodi", 7, "Place"),
    ("Latitude / Shii", 7, "Meal"),
    ("Ateneum free Thu", 6, "Place"),
    ("Cathedral day", 2, "Skip"),
    ("Momotoko", 0, "Skip"),
]
for row in viz:
    ws.append(list(row))
style_header(ws, 3)
paint(ws)
chart = BarChart()
chart.type = "bar"
chart.title = "Helsinki must-scores (Bob-specific, not TripAdvisor)"
chart.x_axis.title = "Score 0–10"
data = Reference(ws, min_col=2, min_row=1, max_row=17)
cats = Reference(ws, min_col=1, min_row=2, max_row=17)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.style = 10
chart.legend = None
chart.height = 12
chart.width = 18
ws.add_chart(chart, "E2")
ws.conditional_formatting.add(
    "B2:B17",
    ColorScaleRule(
        start_type="num", start_value=0, start_color="E8D5D5",
        mid_type="num", mid_value=6, mid_color="FFF3C4",
        end_type="num", end_value=10, end_color="2A6F6F",
    ),
)
autosize(ws, [28, 14, 12])
ws.sheet_properties.tabColor = NAVY

# neighbourhood counts for a second chart
ws["A20"] = "Neighbourhood"
ws["B20"] = "Must_places_on_shortlist"
ws["A20"].fill = header_fill
ws["B20"].fill = header_fill
ws["A20"].font = white_font
ws["B20"].font = white_font
nbh = [
    ("Punavuori / Design District", 8),
    ("Kamppi / Lasipalatsi", 3),
    ("Ullanlinna / Kaartinkaupunki", 4),
    ("Suomenlinna", 1),
    ("Hernesaari", 1),
    ("Kallio", 2),
    ("Töölö", 1),
    ("Kalasatama (out of way)", 1),
]
for i, row in enumerate(nbh, 21):
    ws.cell(i, 1, row[0]).font = body
    ws.cell(i, 2, row[1]).font = body
    ws.cell(i, 1).border = thin
    ws.cell(i, 2).border = thin
chart2 = BarChart()
chart2.type = "col"
chart2.title = "Where the shortlist actually sits"
data2 = Reference(ws, min_col=2, min_row=20, max_row=28)
cats2 = Reference(ws, min_col=1, min_row=21, max_row=28)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.legend = None
chart2.height = 10
chart2.width = 16
ws.add_chart(chart2, "E22")

# ── 10 SOURCES ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("10_Sources")
sh = ["ID", "What", "URL", "Accessed", "Confidence"]
sources = [
    ["S-ekberg", "Ekberg hours + brunch + reservation rules", "https://www.ekberg.fi/", ACCESSED, "High — official"],
    ["S-ekberg-book", "Ekberg table reservation", "https://www.ekberg.fi/en/cafe/table-reservation", ACCESSED, "High — official"],
    ["S-fazer", "Fazer Café Kluuvikatu hours", "https://www.fazer.fi/fazer-cafe/kahvilat/karl-fazer-cafe/", ACCESSED, "High — official café hours; brunch sitting VERIFY"],
    ["S-gk", "Georgian Kitchen hours + phone", "https://www.georgiankitchen.fi/", ACCESSED, "High — official"],
    ["S-rioni", "Rioni Helsinki hours", "https://www.rioni.fi/en-gb/helsinki", ACCESSED, "High — official"],
    ["S-supra", "Georgian Supra address", "https://www.georgiansupra.fi/contact", ACCESSED, "High address; hours secondary"],
    ["S-gaijin", "Gaijin hours + summer lunch closure", "https://www.gaijin.fi/contact", ACCESSED, "High — official"],
    ["S-nolla", "Nolla reservations", "https://restaurantnolla.com/reservations/", ACCESSED, "High booker; hours not on official page"],
    ["S-gron-menu", "Grön summer 2026 menu 188e", "https://www.restaurantgron.com/menu", ACCESSED, "High — official"],
    ["S-gron-book", "Grön DinnerBooking", "https://dinnerbooking.com/fi/en-US/r3904/restaurant-gron", ACCESSED, "High — official booker"],
    ["S-lat25", "Latitude 25 hours", "https://www.latitude25.fi/", ACCESSED, "High — official"],
    ["S-shii", "Shii concept + phone", "https://shii.fi/", ACCESSED, "High concept; hours missing"],
    ["S-seahorse", "Sea Horse summer hours", "https://www.seahorse.fi/en/restaurant-sea-horse/", ACCESSED, "High — official"],
    ["S-kosmos", "Kosmos hours + menu", "https://kosmos.fi/en/our-food/", ACCESSED, "High — official"],
    ["S-elite", "Elite hours", "https://elite.fi/en/front-page/", ACCESSED, "High — official"],
    ["S-bui", "Bui Ramen hours + prices", "https://buiramen.fi/", ACCESSED, "High — official"],
    ["S-levain", "Levain branch hours", "https://www.levain.fi/reservation", ACCESSED, "High — official"],
    ["S-loyly", "Löyly sauna price + rules", "https://www.loylyhelsinki.fi/en/public-sauna", ACCESSED, "High — official"],
    ["S-loyly-h", "Löyly weekly hours", "https://www.loylyhelsinki.fi/en/info", ACCESSED, "High — official"],
    ["S-rex", "Amos Rex hours", "https://amosrex.fi/en/visit-us/", ACCESSED, "High — official"],
    ["S-rex-t", "Amos Rex tickets", "https://amosrex.fi/en/tickets/", ACCESSED, "High — official"],
    ["S-adm", "ADM hours + construction", "https://admuseo.fi/en/visit-us/", ACCESSED, "High — official"],
    ["S-oodi", "Oodi hours 2026", "https://oodihelsinki.fi/en/arrival/", ACCESSED, "High — official"],
    ["S-hsl", "HSL Suomenlinna ferry", "https://www.hsl.fi/en/travelling/visitors/suomenlinna", ACCESSED, "High — official; fare can move"],
    ["S-sl-arr", "Suomenlinna arriving", "https://suomenlinna.fi/en/explore/arriving/", ACCESSED, "High — official"],
    ["S-sl-tour", "English guided tour 2026", "https://suomenlinna.johku.com/en_US/liput/guided-walking-tour-in-english", ACCESSED, "High schedule; inventory unknown"],
    ["S-ateneum", "Ateneum hours + free 20 Aug", "https://ateneum.fi/en/opening-hours-and-tickets/", ACCESSED, "High — official"],
    ["S-allas", "Allas hours", "https://allasseapool.fi/en/", ACCESSED, "High — official"],
    ["S-momo", "Momotoko bankruptcy", "https://www.hs.fi/helsinki/art-2000011230551.html", "2025-05-15", "High — local press, closure status"],
    ["S-katana", "Katana hours (Falstaff)", "https://www.falstaff.com/at/streetfood/katana-ramen-restaurant", ACCESSED, "Medium — secondary, no official hours page"],
]
add_table(ws, sh, sources, [14, 40, 70, 14, 40], GOLD, 28)
for r in range(2, ws.max_row + 1):
    url = ws.cell(r, 3).value
    if url and str(url).startswith("http"):
        href(ws, r, 3, url)

wb.save(OUT)

# CSV dumps for git-diffable tables
import csv

def dump_csv(title):
    sheet = wb[title]
    path = HERE / f"{title}.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            if any(v is not None and str(v).strip() for v in row):
                w.writerow(["" if v is None else v for v in row])
    return path

# only dump the tabular sheets
for name in ["01_Itinerary", "02_Book_tonight", "03_Restaurants", "04_Arrival_brunch",
             "05_Activities", "06_Tours_bookable", "07_Open_matrix", "08_Rain_and_skip",
             "09_Viz_data", "10_Sources"]:
    dump_csv(name)

print("Wrote", OUT)
